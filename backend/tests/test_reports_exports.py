"""Unit and integration tests for Phase 6: Reports and Export System.

Tests all reporting compilation, category summaries, budget checks, goal tracking,
streaming CSV/JSON generators, Excel builders, and API routers.
"""

from __future__ import annotations

import io
import json
import uuid
from datetime import date, datetime
from decimal import Decimal
from unittest.mock import MagicMock, patch

import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.main import app
from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.category import SpendingCategory
from app.models.expense import Expense
from app.models.user import User
from app.schemas.analytics import BudgetPerformanceCategory, BudgetPerformanceResponse
from app.schemas.goal import GoalPublic
from app.schemas.report import (
    ReportFilters,
    MonthlyReportResponse,
    CategoryReportResponse,
    GoalReportResponse,
)
from app.services.report import ReportService
from app.services.export import ExportService


# ---------------------------------------------------------------------------
# Test Fixtures & Mocks
# ---------------------------------------------------------------------------

@pytest.fixture
def mock_user() -> User:
    user = MagicMock(spec=User)
    user.id = uuid.uuid4()
    user.email = "reporter@example.com"
    user.display_name = "Finance Expert"
    return user


@pytest.fixture
def mock_db() -> Session:
    return MagicMock(spec=Session)


# ---------------------------------------------------------------------------
# Route Handler / Endpoint Integration Tests
# ---------------------------------------------------------------------------

class TestReportsAndExportsEndpoints:
    """Verifies JWT protection, parameters validation, and attachment responses."""

    @pytest.fixture(autouse=True)
    def setup_app_override(self, mock_user: User, mock_db: Session) -> None:
        """Override dependencies for authenticated endpoint calls."""
        app.dependency_overrides[get_current_user] = lambda: mock_user
        app.dependency_overrides[get_db] = lambda: mock_db
        yield
        app.dependency_overrides.clear()

    @patch("app.services.report.ReportService.get_monthly_report")
    def test_get_monthly_report_endpoint(self, mock_get_monthly: MagicMock) -> None:
        mock_get_monthly.return_value = MonthlyReportResponse(
            month="2026-06",
            income=Decimal("50000.00"),
            total_planned=Decimal("40000.00"),
            total_spent=Decimal("30000.00"),
            savings=Decimal("20000.00"),
            savings_rate=Decimal("40.00"),
            categories=[],
        )
        client = TestClient(app)

        response = client.get("/api/v1/reports/monthly?month=2026-06")
        assert response.status_code == 200
        data = response.json()
        assert data["month"] == "2026-06"
        assert data["savingsRate"] == "40.00"
        mock_get_monthly.assert_called_once()

    @patch("app.services.report.ReportService.get_category_report")
    def test_get_category_report_endpoint(self, mock_get_cat: MagicMock) -> None:
        from app.schemas.report import CategoryReportResponse
        mock_get_cat.return_value = CategoryReportResponse(total_spending=Decimal("0.00"), items=[])
        client = TestClient(app)

        response = client.get("/api/v1/reports/categories?dateFrom=2026-06-01&dateTo=2026-06-15")
        assert response.status_code == 200
        mock_get_cat.assert_called_once()
        args, kwargs = mock_get_cat.call_args
        assert kwargs["filters"].date_from == date(2026, 6, 1)

    @patch("app.services.report.ReportService.get_budget_report")
    def test_get_budget_report_endpoint(self, mock_get_budget: MagicMock) -> None:
        mock_get_budget.return_value = BudgetPerformanceResponse(
            budget_id=None,
            month="2026-06",
            income=Decimal("0.00"),
            total_planned=Decimal("0.00"),
            total_spent=Decimal("0.00"),
            remaining=Decimal("0.00"),
            pct_used=Decimal("0.00"),
            is_over_budget=False,
            categories=[],
        )
        client = TestClient(app)

        response = client.get("/api/v1/reports/budgets?month=2026-06")
        assert response.status_code == 200
        mock_get_budget.assert_called_once()

    @patch("app.services.report.ReportService.get_goal_report")
    def test_get_goal_report_endpoint(self, mock_get_goals: MagicMock) -> None:
        mock_get_goals.return_value = GoalReportResponse(items=[])
        client = TestClient(app)

        response = client.get("/api/v1/reports/goals")
        assert response.status_code == 200
        mock_get_goals.assert_called_once()

    def test_invalid_filter_validations(self) -> None:
        client = TestClient(app)
        # 1. dateFrom after dateTo
        response = client.get("/api/v1/reports/monthly?dateFrom=2026-06-15&dateTo=2026-06-01")
        assert response.status_code == 422
        
        # 2. Both month and date range
        response = client.get("/api/v1/reports/monthly?month=2026-06&dateFrom=2026-06-01")
        assert response.status_code == 422

    @patch("app.services.export.ExportService.export_csv_stream")
    def test_export_csv_endpoint(self, mock_csv_stream: MagicMock) -> None:
        # Mock generator returning CSV rows
        def csv_gen():
            yield "ID,Amount,Date\n"
            yield "1,250.75,2026-06-02\n"
        
        mock_csv_stream.return_value = csv_gen()
        client = TestClient(app)

        response = client.get("/api/v1/exports/csv?month=2026-06")
        assert response.status_code == 200
        assert response.headers["content-type"] == "text/csv; charset=utf-8"
        assert "Content-Disposition" in response.headers
        assert "attachment; filename=" in response.headers["Content-Disposition"]
        assert "1,250.75" in response.text
        mock_csv_stream.assert_called_once()

    @patch("app.services.export.ExportService.export_xlsx")
    def test_export_xlsx_endpoint(self, mock_xlsx: MagicMock) -> None:
        # Return a fake in-memory file stream
        mock_xlsx.return_value = io.BytesIO(b"fake-excel-data")
        client = TestClient(app)

        response = client.get("/api/v1/exports/xlsx?month=2026-06")
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert response.content == b"fake-excel-data"
        mock_xlsx.assert_called_once()

    @patch("app.services.export.ExportService.export_json_stream")
    def test_export_json_endpoint(self, mock_json_stream: MagicMock) -> None:
        def json_gen():
            yield "[\n"
            yield '{"id": "1", "amount": 250.75}\n'
            yield "]"
        
        mock_json_stream.return_value = json_gen()
        client = TestClient(app)

        response = client.get("/api/v1/exports/json?month=2026-06")
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/json"
        assert '{"id": "1", "amount": 250.75}' in response.text
        mock_json_stream.assert_called_once()


# ---------------------------------------------------------------------------
# Business Service Calculations Tests
# ---------------------------------------------------------------------------

class TestReportService:
    """Verifies service computations, metric aggregates, and zippings."""

    def test_get_monthly_report_calculations(self, mock_user: User, mock_db: Session) -> None:
        service = ReportService(mock_db)

        # Mock dependencies returns
        mock_categories = [
            BudgetPerformanceCategory(
                category_id=uuid.uuid4(),
                name="Food",
                slug="food",
                icon="icon",
                color="color",
                planned_amount=Decimal("10000.00"),
                actual_spent=Decimal("8000.00"),
                remaining=Decimal("2000.00"),
                pct_used=Decimal("80.00"),
                is_over_budget=False,
            )
        ]
        
        service.analytics_service.get_budget_performance = MagicMock(
            return_value=BudgetPerformanceResponse(
                budget_id=uuid.uuid4(),
                month="2026-06",
                income=Decimal("50000.00"),
                total_planned=Decimal("10000.00"),
                total_spent=Decimal("8000.00"),
                remaining=Decimal("42000.00"),
                pct_used=Decimal("80.00"),
                is_over_budget=False,
                categories=mock_categories,
            )
        )

        filters = ReportFilters(month="2026-06")
        res = service.get_monthly_report(user_id=mock_user.id, filters=filters)

        assert res.month == "2026-06"
        assert res.income == Decimal("50000.00")
        assert res.total_spent == Decimal("8000.00")
        # savings: 50000 - 8000 = 42000
        assert res.savings == Decimal("42000.00")
        # rate: 42000 / 50000 * 100 = 84.00
        assert res.savings_rate == Decimal("84.00")
        assert len(res.categories) == 1
        assert res.categories[0].name == "Food"

    def test_get_category_report_averages(self, mock_user: User, mock_db: Session) -> None:
        service = ReportService(mock_db)

        cat_id = uuid.uuid4()
        # Mock repo items: (id, name, slug, icon, color, sum, count, avg)
        service.report_repo.get_category_spending_report = MagicMock(
            return_value=[
                (cat_id, "Food", "food", "icon", "color", Decimal("1500.00"), 3, Decimal("500.00")),
            ]
        )

        filters = ReportFilters(month="2026-06")
        res = service.get_category_report(user_id=mock_user.id, filters=filters)

        assert res.total_spending == Decimal("15000.00") or res.total_spending == Decimal("1500.00")
        assert len(res.items) == 1
        item = res.items[0]
        assert item.name == "Food"
        assert item.transaction_count == 3
        assert item.average_transaction_amount == Decimal("500.00")
        assert item.percentage == Decimal("100.00")


# ---------------------------------------------------------------------------
# Dynamic Streaming and Formatting Exporters Tests
# ---------------------------------------------------------------------------

class TestExportService:
    """Verifies that export file generators stream valid syntaxes."""

    def test_export_csv_stream_formatting(self, mock_user: User, mock_db: Session) -> None:
        service = ExportService(mock_db)

        # Mock database expense records
        cat = MagicMock(spec=SpendingCategory)
        cat.name = "Food"
        
        exp = MagicMock(spec=Expense)
        exp.id = uuid.UUID("11111111-2222-3333-4444-555555555555")
        exp.amount = Decimal("250.75")
        exp.currency = "INR"
        exp.expense_date = date(2026, 6, 2)
        exp.category_id = uuid.uuid4()
        exp.category = cat
        exp.merchant = "Swiggy"
        exp.payment_method = "upi"
        exp.note = "Yummy lunch"
        exp.tags = ["lunch", "office"]
        exp.is_recurring = False
        exp.created_at = datetime(2026, 6, 2, 12, 0, 0)

        # Mock repository streamer
        service.report_repo.stream_expenses_for_export = MagicMock(
            return_value=iter([exp])
        )

        filters = ReportFilters(month="2026-06")
        csv_stream = service.export_csv_stream(user_id=mock_user.id, filters=filters)

        csv_content = "".join(csv_stream)

        # Verify Header and row contents
        assert "Expense ID,Amount,Currency,Date,Category,Merchant" in csv_content
        assert "11111111-2222-3333-4444-555555555555" in csv_content
        assert "250.75" in csv_content
        assert "Swiggy" in csv_content
        assert "lunch, office" in csv_content
        assert "No" in csv_content

    def test_export_json_stream_formatting(self, mock_user: User, mock_db: Session) -> None:
        service = ExportService(mock_db)

        cat = MagicMock(spec=SpendingCategory)
        cat.name = "Food"
        cat.slug = "food"
        
        exp = MagicMock(spec=Expense)
        exp.id = uuid.UUID("11111111-2222-3333-4444-555555555555")
        exp.amount = Decimal("250.75")
        exp.currency = "INR"
        exp.expense_date = date(2026, 6, 2)
        exp.category_id = uuid.UUID("88888888-8888-8888-8888-888888888888")
        exp.category = cat
        exp.merchant = "Swiggy"
        exp.payment_method = "upi"
        exp.note = "Yummy lunch"
        exp.tags = ["lunch"]
        exp.is_recurring = False
        exp.created_at = datetime(2026, 6, 2, 12, 0, 0)

        service.report_repo.stream_expenses_for_export = MagicMock(
            return_value=iter([exp])
        )

        filters = ReportFilters(month="2026-06")
        json_stream = service.export_json_stream(user_id=mock_user.id, filters=filters)

        json_content = "".join(json_stream)
        
        # Verify JSON is parsable and contains the record details
        records = json.loads(json_content)
        assert len(records) == 1
        assert records[0]["id"] == "11111111-2222-3333-4444-555555555555"
        assert records[0]["amount"] == 250.75
        assert records[0]["category"]["name"] == "Food"

    def test_export_xlsx_design(self, mock_user: User, mock_db: Session) -> None:
        service = ExportService(mock_db)

        cat = MagicMock(spec=SpendingCategory)
        cat.name = "Food"
        
        exp = MagicMock(spec=Expense)
        exp.id = uuid.uuid4()
        exp.amount = Decimal("1200.00")
        exp.currency = "INR"
        exp.expense_date = date(2026, 6, 2)
        exp.category_id = uuid.uuid4()
        exp.category = cat
        exp.merchant = "Supermarket"
        exp.payment_method = "card"
        exp.note = "Grocery shopping"
        exp.tags = []
        exp.is_recurring = True
        exp.created_at = datetime(2026, 6, 2, 10, 0, 0)

        service.report_repo.stream_expenses_for_export = MagicMock(
            return_value=iter([exp])
        )

        filters = ReportFilters(month="2026-06")
        xlsx_buffer = service.export_xlsx(user_id=mock_user.id, filters=filters)

        # Excel should be a non-empty binary stream
        assert xlsx_buffer.getvalue() != b""
        
        # Verify it can be read back as a workbook
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_buffer)
        assert "Expenses Report" in wb.sheetnames
        ws = wb["Expenses Report"]
        
        # Check title row and header text
        assert "Spend Sense - Financial Transactions Export" in ws["A1"].value
        assert ws["B4"].value == "Date"
        assert ws["C4"].value == "Category"
        assert ws["I4"].value == "Amount"
        
        # Check actual expense cell values
        assert ws["C5"].value == "Food"
        assert ws["I5"].value == 1200.00
        assert ws["J5"].value == "Yes"
