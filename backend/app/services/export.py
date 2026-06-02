"""Service for formatting and exporting financial data.

Handles CSV streaming, JSON streaming, and styled in-memory Excel generation.
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import date
from decimal import Decimal
from typing import Generator, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.repositories.report import ReportRepository
from app.schemas.report import ReportFilters


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _resolve_dates(filters: ReportFilters) -> tuple[Optional[date], Optional[date]]:
    """Extract start and end dates from report filters."""
    from app.services.report import _month_date_range

    if filters.month:
        return _month_date_range(filters.month)
    return filters.date_from, filters.date_to


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------

class ExportService:
    """Manages parsing and streaming formatting for file exports."""

    def __init__(self, db: Session) -> None:
        self.db = db
        self.report_repo = ReportRepository(db)

    def export_csv_stream(
        self,
        user_id: uuid.UUID,
        filters: ReportFilters,
    ) -> Generator[str, None, None]:
        """Stream user expenses as a memory-safe CSV file."""
        start_date, end_date = _resolve_dates(filters)

        output = io.StringIO()
        writer = csv.writer(output)

        # 1. Header row
        writer.writerow([
            "Expense ID",
            "Amount",
            "Currency",
            "Date",
            "Category",
            "Merchant",
            "Payment Method",
            "Note",
            "Tags",
            "Is Recurring",
            "Created At",
        ])
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        # 2. Dynamic database row streaming
        expense_generator = self.report_repo.stream_expenses_for_export(
            user_id=user_id,
            start_date=start_date,
            end_date=end_date,
            category_id=filters.category_id,
        )

        for expense in expense_generator:
            writer.writerow([
                str(expense.id),
                f"{expense.amount:.2f}",
                expense.currency,
                expense.expense_date.isoformat(),
                expense.category.name if expense.category else "Uncategorized",
                expense.merchant or "",
                expense.payment_method or "",
                expense.note,
                ", ".join(expense.tags) if expense.tags else "",
                "Yes" if expense.is_recurring else "No",
                expense.created_at.isoformat() if expense.created_at else "",
            ])
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    def export_json_stream(
        self,
        user_id: uuid.UUID,
        filters: ReportFilters,
    ) -> Generator[str, None, None]:
        """Stream user expenses as a memory-safe JSON array."""
        start_date, end_date = _resolve_dates(filters)

        # Yield open bracket
        yield "[\n"

        expense_generator = self.report_repo.stream_expenses_for_export(
            user_id=user_id,
            start_date=start_date,
            end_date=end_date,
            category_id=filters.category_id,
        )

        first = True
        for expense in expense_generator:
            if not first:
                yield ",\n"
            first = False

            record = {
                "id": str(expense.id),
                "amount": float(expense.amount),
                "currency": expense.currency,
                "expenseDate": expense.expense_date.isoformat(),
                "category": {
                    "id": str(expense.category_id),
                    "name": expense.category.name if expense.category else "Uncategorized",
                    "slug": expense.category.slug if expense.category else "others",
                } if expense.category else None,
                "merchant": expense.merchant,
                "paymentMethod": expense.payment_method,
                "note": expense.note,
                "tags": expense.tags,
                "isRecurring": expense.is_recurring,
                "createdAt": expense.created_at.isoformat() if expense.created_at else None,
            }
            yield json.dumps(record, indent=2)

        # Yield close bracket
        yield "\n]"

    def export_xlsx(
        self,
        user_id: uuid.UUID,
        filters: ReportFilters,
    ) -> io.BytesIO:
        """Generate a beautifully formatted and styled in-memory Excel sheet."""
        start_date, end_date = _resolve_dates(filters)

        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses Report"
        ws.views.sheetView[0].showGridLines = True

        # Styles definition (Premium Slate Theme)
        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # deep slate
        
        row_font = Font(name=font_family, size=10)
        cardinal_font = Font(name=font_family, size=10, bold=True)
        
        thin_side = Side(border_style="thin", color="CBD5E1") # gray border
        double_side = Side(border_style="double", color="475569")
        
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_total = Border(top=thin_side, bottom=double_side)

        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        # 1. Sheet Title block
        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value = "Spend Sense - Financial Transactions Export"
        title_cell.font = Font(name=font_family, size=16, bold=True, color="1E293B")
        title_cell.alignment = align_left
        ws.row_dimensions[1].height = 35

        # 2. Date ranges info
        ws.merge_cells("A2:K2")
        meta_cell = ws["A2"]
        range_str = f"Date Range: {start_date.isoformat()} to {end_date.isoformat()}" if start_date and end_date else "Date Range: Full History"
        meta_cell.value = f"Generated on: {date.today().isoformat()}  |  {range_str}"
        meta_cell.font = Font(name=font_family, size=9, italic=True, color="64748B")
        meta_cell.alignment = align_left
        ws.row_dimensions[2].height = 20

        # Empty spacing row
        ws.row_dimensions[3].height = 15

        # 3. Header Setup
        headers = [
            "Expense ID",
            "Date",
            "Category",
            "Merchant",
            "Payment Method",
            "Note",
            "Tags",
            "Currency",
            "Amount",
            "Is Recurring",
            "Created At",
        ]

        ws.row_dimensions[4].height = 26
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = Border(top=thin_side, bottom=thin_side)

        # 4. Fetch and fill data
        expense_generator = self.report_repo.stream_expenses_for_export(
            user_id=user_id,
            start_date=start_date,
            end_date=end_date,
            category_id=filters.category_id,
        )

        current_row = 5
        total_amount = Decimal("0.00")

        for expense in expense_generator:
            ws.row_dimensions[current_row].height = 20
            
            # Format and set values
            c_id = ws.cell(row=current_row, column=1, value=str(expense.id))
            c_date = ws.cell(row=current_row, column=2, value=expense.expense_date)
            c_cat = ws.cell(row=current_row, column=3, value=expense.category.name if expense.category else "Uncategorized")
            c_merch = ws.cell(row=current_row, column=4, value=expense.merchant or "")
            c_pay = ws.cell(row=current_row, column=5, value=expense.payment_method or "")
            c_note = ws.cell(row=current_row, column=6, value=expense.note)
            c_tags = ws.cell(row=current_row, column=7, value=", ".join(expense.tags) if expense.tags else "")
            c_curr = ws.cell(row=current_row, column=8, value=expense.currency)
            c_amt = ws.cell(row=current_row, column=9, value=float(expense.amount))
            c_rec = ws.cell(row=current_row, column=10, value="Yes" if expense.is_recurring else "No")
            c_create = ws.cell(row=current_row, column=11, value=expense.created_at.strftime("%Y-%m-%d %H:%M:%S") if expense.created_at else "")

            # Mathematical tracking
            total_amount += expense.amount

            # Styling row cells
            c_id.alignment = align_center
            c_date.alignment = align_center
            c_date.number_format = "yyyy-mm-dd"
            c_cat.alignment = align_left
            c_merch.alignment = align_left
            c_pay.alignment = align_center
            c_note.alignment = align_left
            c_tags.alignment = align_left
            c_curr.alignment = align_center
            c_amt.alignment = align_right
            c_amt.number_format = "#,##0.00" # decimal formatting
            c_rec.alignment = align_center
            c_create.alignment = align_center

            for col in range(1, 12):
                cell = ws.cell(row=current_row, column=col)
                cell.font = row_font
                cell.border = border_all

            current_row += 1

        # 5. Total Row
        ws.row_dimensions[current_row].height = 24
        
        ws.cell(row=current_row, column=7, value="Total Spending:").font = cardinal_font
        ws.cell(row=current_row, column=7).alignment = align_right
        
        tot_curr = ws.cell(row=current_row, column=8, value="INR") # Default/Main currency info
        tot_curr.font = cardinal_font
        tot_curr.alignment = align_center
        tot_curr.border = border_total

        tot_amt = ws.cell(row=current_row, column=9, value=float(total_amount))
        tot_amt.font = cardinal_font
        tot_amt.alignment = align_right
        tot_amt.number_format = "#,##0.00"
        tot_amt.border = border_total

        # Set borders on total descriptive label
        ws.cell(row=current_row, column=7).border = Border(top=thin_side, bottom=double_side)

        # 6. Auto-fit Column Widths cleanly
        for col in ws.columns:
            # Avoid measuring titles in merged row 1 and 2
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                # ignore rows 1 & 2 for layout width checks
                if cell.row in (1, 2, 3):
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            
            # Pad
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Specific custom spacing overrides
        ws.column_dimensions["A"].width = 38 # ID column
        ws.column_dimensions["F"].width = 25 # Note column
        ws.column_dimensions["K"].width = 20 # Timestamp column

        # 7. Write workbook to memory stream
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        return xlsx_buffer
